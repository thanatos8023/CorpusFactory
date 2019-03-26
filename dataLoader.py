#-*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import glob
from werkzeug import filesystem

# Load Question to Intention file
# Frame:
#    Intention     Question                 SlotFilling
#   Intention_A   Question_A   <SLOTNAME1>:VALUE1,<SLOTNAME2>:VALUE2,...
#   Intention_B   Question_B   <SLOTNAME1>:VALUE1,<SLOTNAME2>:VALUE2,...
#       ...           ...                       ...
def load_q2i(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        raw = f.read().split('\n')
    
    # result: 
    # [
    #    {'intention': Intention_A, 'question': Question_A, 'slots': [(<SLOTNAME1>,VALUE1), (<SLOTNAME2>,VALUE2),...]},
    #    {'intention': Intention_B, 'question': Question_B, 'slots': [(<SLOTNAME1>,VALUE1), (<SLOTNAME2>,VALUE2),...]},
    #         ...
    # ]
    result = list()
    for line in raw:
        temp = line.split('\t')
        if len(temp) < 3:
            result.append(
                {
                    'intention': temp[0].strip(),
                    'question': temp[1].strip()
                }
            )
        else:
            result.append(
                {
                    'intention': temp[0].strip(),
                    'question': temp[1].strip(),
                    'slots': slot_string_transform(temp[2].strip())
                }
            )
    
    return result

# Input form: <SLOTNAME1>:VALUE1,<SLOTNAME2>:VALUE2,...
# Output form: [(<SLOTNAME1>, VALUE1), (<SLOTNAME2>, VALUE2), ...]
def slot_string_transform(slot_string):
    result = list()
    slot_list = slot_string.split(',')
    for slot in slot_list:
        slotname = slot.split(':')[0].strip()
        slotvalue = slot.split(':')[1].strip()
        result.append((slotname, slotvalue))
    return result

## Load excel 
def excel_to_rawcorpus(path, q2i_list):
    result = ''
    # Path is directory
    if os.path.isdir(path):
        filelist = glob.glob('work/*.xls*')
        for filename in filelist:
            wb = load_workbook(filename=filename)
            sheet_ranges = wb['IOT']
            
            for i in range(10, 99999, 13):
                cell_idx = 'C%d' % i
                if not sheet_ranges[cell_idx].value: break
                question = sheet_ranges[cell_idx].value # Question
                intention = q2i(q2i_list, question)
                for j in range(10):
                    sent_idx = 'C%d' % (i+3+j)
                    text = sheet_ranges[sent_idx].value # Sentence
                    if not text: break
                    text = slot_fill(q2i_list, intention, text)
                    result += '%s\t%s\n' % (text, intention)
    # Path is filename
    else:
        wb = load_workbook(filename=path)
        sheet_ranges = wb['IOT']

        for i in range(10, 99999, 13):
            cell_idx = 'C%d' % i
            if not sheet_ranges[cell_idx].value: break
            question = sheet_ranges[cell_idx].value # Question
            intention = q2i(q2i_list, question)
            for j in range(10):
                sent_idx = 'C%d' % (i+3+j)
                text = sheet_ranges[sent_idx].value # Sentence
                text = slot_fill(q2i_list, intention, text)
                result += '%s\t%s\n' % (text, intention)      
    return result

## Get intention using question 
def q2i(q2i_list, question):
    intention = ''
    for obj in q2i_list:
        if obj['question'] == question:
            intention = obj['intention']
            break
    return intention

## Slot filling
def slot_fill(q2i, intention, text):
    for obj in q2i:
        if obj['intention'] == intention and len(obj) > 2:
            slot_list = obj['slots']
            break
        else:
            slot_list = []
    for slot in slot_list:
        text = text.replace(slot[1], slot[0])
    return text

# Save the corpus raw 
def save_raw(textform, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(textform)
    print("SAVING...", filename, 'corpus raw saved!')

# Load the corpus raw
def load_raw(filename):
    with open(filename, 'rb') as f:
        data = f.read()
        try:
            data = data.decode('utf-8')
        except UnicodeDecodeError:
            data = data.decode('euc-kr')
        raw = data.split('\n')
    result = dict()
    for line in raw:
        # End check
        if not line: continue
            
        intention = line.split('\t')[1]
        text = line.split('\t')[0]
        if intention in result.keys():
            result[intention].append(text)
        else:
            result[intention] = [text]
            
    return result

def make_raw(q2i_filename, excel_filepath):
    print('[[[ Step 1 ]]]\tLoading Mapping (Question to Intention) file ')
    q2i_list = load_q2i(q2i_filename)
    print("Result:", q2i_list[:5], '\n')

    print('[[[ Step 2 ]]]\tLoading and Converting (Excel to Python list object) file')
    corpusraw = excel_to_rawcorpus(excel_filepath, q2i_list)
    print("Result:", corpusraw[:100], '\n')

    print('[[[ Step 3 ]]]\tSaving raw curpus file (txt)')
    save_raw(corpusraw, 'raw.txt')
    print('Result:', 'raw.txt' + " SAVED in", os.getcwd())

    print('[[[ Step 4 ]]]\tLoading raw corpus file (txt) using load_raw() function')
    raw = load_raw('raw.txt')
    print('Result:', raw['ac_set'][:5], '\n')

# Test code
if __name__ == '__main__':
    make_raw("q2i.txt", "device_Corpus collection_180710_Test (1).xlsx")